import pandas as pd
import openpyxl
from datetime import datetime
from functools import wraps
from flask import session, redirect, url_for, flash



# Función para actualizar el estado de un usuario
def update_user(user_id, data):
    df = get_users()
    df.loc[df['id'] == user_id, ['tipo_de_documento', 'nuip', 'nombres', 'apellidos', 'fecha_nacimiento', 'correo', 'usuario', 'contraseña', 'telefono', 'rol', 'estado_usuario']] = [
        data['tipo_de_documento'], data['nuip'], data['nombres'], data['apellidos'], data['fecha_nacimiento'], data['correo'], data['usuario'], data['contraseña'], data['telefono'], data['rol'], data['estado_usuario']
    ]
    df.to_excel('usuarios.xlsx', index=False)

# Función para eliminar un usuario
def delete_user(user_id):
    df = get_users()
    df = df[df['id'] != user_id]
    df.to_excel('usuarios.xlsx', index=False)

# Función para cambiar el estado de un usuario
def toggle_user_status(user_id, new_status):
    valid_status = ['activo', 'inactivo', 'suspendido', 'baneado', 'sancionado', 'vacaciones', 'vacio', 'retirado']
    
    if new_status not in valid_status:
        raise Exception("Estado no válido")
    
    df = get_users()
    df.loc[df['id'] == user_id, 'estado_usuario'] = new_status
    df.to_excel('usuarios.xlsx', index=False)






def role_required(roles):
    """
    Decorador que limita el acceso a una vista según el rol del usuario.
    """
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            # Verificar si el usuario está autenticado y si tiene el rol adecuado
            if 'usuario' not in session or session['rol'] not in roles:
                flash('Acceso no autorizado', 'danger')
                return redirect(url_for('auth.login'))  # Redirige al login si no tiene el rol adecuado
            return f(*args, **kwargs)
        return wrapped
    return decorator

# Leer el archivo de usuarios
def get_users():
    """Obtiene todos los usuarios desde el archivo Excel."""
    try:
        df = pd.read_excel('usuarios.xlsx')
        return df
    except FileNotFoundError:
        raise Exception("Archivo de usuarios no encontrado.")

# Filtrar usuarios por rol
def get_users_by_role(role):
    """Filtra los usuarios por su rol."""
    df = get_users()
    return df[df['rol'] == role]

# Obtener el rol de un usuario
def get_user_role(username):
    try:
        # Leer el archivo Excel
        df = pd.read_excel('usuarios.xlsx')

        # Depurar: Mostrar las columnas del DataFrame
        print("Columnas del DataFrame:", df.columns)

        # Filtrar el usuario por el nombre de usuario
        user = df[df['usuario'] == username]

        if not user.empty:
            return user.iloc[0]['rol']  # Devuelve el rol del primer usuario encontrado
        else:
            raise Exception(f"Usuario '{username}' no encontrado.")

    except Exception as e:
        raise Exception(f"Error al obtener el rol del usuario: {e}")


def load_users(file_path='usuarios.xlsx'):
    """Carga los usuarios desde un archivo Excel."""
    try:
        return pd.read_excel(file_path)
    except FileNotFoundError:
        raise Exception("Archivo de usuarios no encontrado.")

def validate_user(username, password):
    """Valida si un usuario y contraseña son correctos."""
    excel_path = 'usuarios.xlsx'

    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        # Recorrer las filas y verificar usuario y contraseña
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignorar encabezados
            excel_username = row[6]  # Columna de "Usuario"
            excel_password = row[7]  # Columna de "Contraseña"
            if excel_username == username and excel_password == password:
                return True  # Usuario válido
        return False  # Usuario o contraseña incorrectos
    except FileNotFoundError:
        raise Exception("El archivo de usuarios no existe. Registra al menos un usuario.")
    except Exception as e:
        raise Exception(f"Error al validar usuario: {e}")

def register_user(data):
    """Registra un nuevo usuario en el archivo Excel."""
    excel_path = 'usuarios.xlsx'

    try:
        # Abre el archivo Excel o crea uno nuevo si no existe
        try:
            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            # Escribe los encabezados
            sheet.append([
                'Tipo de Documento', 'NUIP', 'Nombres', 'Apellidos',
                'Fecha de Nacimiento', 'Correo Electrónico', 'Usuario',
                'Contraseña', 'Teléfono', 'Rol', 'Fecha y Hora de Registro'
            ])

        # Agrega los datos del usuario
        sheet.append([
            data['tipo_de_documento'],
            data['nuip'],
            data['nombres'],
            data['apellidos'],
            data['fecha_nacimiento'],
            data['correo'],
            data['usuario'],
            data['contraseña'],  # En producción, encripta esta información
            data['telefono'],
            data['rol'],  # Agregar el rol
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # Fecha y hora actual
        ])

        # Guarda el archivo Excel
        workbook.save(excel_path)
    except Exception as e:
        raise Exception(f"Error al registrar usuario: {e}")
